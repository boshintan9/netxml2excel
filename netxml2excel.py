"""A module for converting multiple Kismet netxml files to XLSX Excel
spreadsheets.
Original script by 'Meatballs'.
Ref: https://github.com/Meatballs1/NetXML-to-CSV
Additional code by 'oldsea'
Ref: https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
Adapted for excel by Brett.
"""


import argparse
import os
import string
import sys
from pathlib import Path
from xml.etree import ElementTree as ETree

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def main():
    """The mainline of the program. Collects arguments and runs accordingly."""

    parser = argparse.ArgumentParser(description='NetXML to XSLX')
    parser.add_argument(
        'input', nargs='+',
        help='One or more netxml files, space separated.'
    )
    parser.add_argument(
        '-o', '--output', metavar='output', default='netxml.xlsx',
        help='Output file path. Defaults to "./netxml.xslx" if left blank.'
    )
    parser.add_argument(
        '-d', '--dir', action='store_true', default=False,
        help=(
            'Use when supplying a directory of netxml files instead of direct '
            'file references.'
        )
    )
    args = parser.parse_args()

    out_path = Path(args.output).resolve()
    if out_path.is_file():
        print(f'{out_path} already exists. Quitting...')
        return

    input_paths = []
    for i in args.input:
        try:
            input_path = Path(i).resolve()
        except Exception as e:
            print(
                f'Could not resolve the file path for {i}. It will be skipped'
            )
        if args.dir and input_path.is_dir():
            netxml_glob = input_path.glob('*.netxml')
            for path in netxml_glob:
                input_paths.append(path)
        elif not input_path.is_file():
            thing = 'directory' if args.dir else 'file'
            print(f'{i} is not a {thing}. Skipping...')
            continue
        else:
            input_paths.append(input_path)
    # Create the in-memory Excel Workbook
    wb = Workbook()
    networks_sheet = wb.active
    networks_sheet.title = 'Wireless Networks'
    # Add the title row for the networks WorkSheet
    networks_sheet.append(
        [
            'BSSID', 'ESSID', 'Hidden', 'Channel', 'Signal Strength', 'Open',
            'WEP', 'WPA', 'WPA2', 'WPS', 'Auth', 'TKIP', 'AES', 'Manufacturer',
            'No. Clients', 'Latitude', 'Longitude'
        ]
    )
    clients_sheet = wb.create_sheet(title='Clients')
    # Add the title row for the clients WorkSheet
    clients_sheet.append(
        ['MAC', 'Manufacturer', 'Signal Strength', 'BSSID', 'ESSID']
    )
    networks_list = []
    clients_list = []
    for input_path in input_paths:
        _parse_netxml(input_path, networks_list, clients_list)
    # Ensure no duplicates end up in the spreadsheets.
    unique_networks = set(networks_list)
    unique_clients = set(clients_list)

    # Add the results of all files to the spreadsheets
    for row in unique_networks:
        networks_sheet.append(row)

    # Add total row
    for row in unique_clients:
        clients_sheet.append(row)
    # Turn the resulting tables in to Excel "Tables"
    _create_table(networks_sheet, 'Networks')
    _create_table(clients_sheet, 'Clients')
    # Create totals WorkSheet
    totals_sheet = wb.create_sheet(title='Totals')
    _populate_totals(totals_sheet, networks_sheet)
    _create_table(totals_sheet, 'Totals')

    wb.save(str(out_path))


def _parse_netxml(input_path, n_list, c_list):
    """Takes a Path object pointing to a netxml file and two openpyxl.Workbook
    WorkSheet objects for the networks and clients.
    """

    try:
        # Read in the netxml file. If it isn't valid it will error out.
        doc = ETree.parse(input_path.open())
    except:
        print(
            f"[-] Unable to open input file: {input_path}. The following "
            "error occurred: \n"
        )
        raise

    for network in doc.getiterator("wireless-network"):
        net_type = network.attrib["type"]
        channel = network.find('channel').text
        bssid = network.find('BSSID').text
        manuf = network.find('manuf').text

        if net_type == "probe" or channel == "0":
            continue

        OPEN = 0
        WEP = 0
        WPA = 0
        WPA2 = 0
        AUTH = 'N/A'
        TKIP = 0
        AES = 0

        ssid = network.find('SSID')
        for e in ssid.findall('encryption'):
            if e.text.startswith("WEP"):
                WEP = 1
                break
            elif e.text.startswith("WPA"):
                if e.text.endswith("PSK"):
                    AUTH = "PSK"
                elif e.text.endswith("AES-CCM"):
                    AES = 1
                elif e.text.endswith("TKIP"):
                    TKIP = 1
            elif e.text == "None":
                OPEN = 1

        essid_text = ""
        if ssid is not None:
            wpa_ver = ssid.find('wpa-version')
            # The wpa-version element usually looks like this: 'WPA+WPA2'
            vers = wpa_ver.text.split('+') if wpa_ver is not None else ''
            WPA = 1 if 'WPA' in vers else 0
            WPA2 = 1 if 'WPA2' in vers else 0
            essid = ssid.find('essid')
            essid_text = essid.text
            cloaked = 1 if essid.attrib['cloaked'] == 'true' else 0
            WPS = ssid.find('wps').text

        power = network.find('snr-info')
        dbm = ""
        if power is not None:
            dbm = power.find('max_signal_dbm').text

        if int(dbm) > 1:
            dbm = power.find('last_signal_dbm').text

        if int(dbm) > 1:
            dbm = power.find('min_signal_dbm').text


        gps = network.find('gps-info')
        lat, lon = '', ''
        if gps is not None:
            lat = network.find('gps-info').find('avg-lat').text
            lon = network.find('gps-info').find('avg-lon').text

        c_count = _add_associated_clients(network, bssid, essid_text, c_list)

        n_list.append(
            (
                bssid, essid_text, cloaked, channel, dbm, OPEN, WEP, WPA,
                WPA2, WPS, AUTH, TKIP, AES, manuf, c_count, lat, lon,
            )
        )


def _add_associated_clients(network, bssid, essid_text, c_list):
    """Summarise all associated clients with the passed in "network".
    Returns the number of clients associated with a Network.
    """

    clients = network.getiterator('wireless-client')
    c_count = 0

    if clients is None:
        return c_count
    for client in clients:
        mac = client.find('client-mac')
        if mac is None:
            continue
        client_mac = mac.text
        snr = client.find('snr-info')
        if snr is None:
            continue
        power = client.find('snr-info').find('max_signal_dbm')
        if power is None:
            continue
        client_power = power.text
        manuf = client.find('client-manuf')
        if manuf is None:
            client_manf = ''
        else:
            client_manf = manuf.text
        c_list.append(
            (
                client_mac, client_manf, client_power, bssid, essid_text
            )
        )
        c_count += 1

    return c_count


def _create_table(sheet, title, style='TableStyleMedium9'):
    """Takes a worksheet assuming there is only one table of data already
    present. Creates an 'Excel Table' from the max_column and max_row range
    giving it the passed in title.

    A default style is also applied and the column widths adjusted.
    """
    # Code originally from:
    # https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 3 if max_length < 10 else max_length
        sheet.column_dimensions[column].width = adjusted_width
    c = sheet.max_column - 1
    r = sheet.max_row - 1
    coords = _indices_to_coords(c,r)
    table = Table(displayName=title, ref=f"A1:{coords['coord']}")
    style = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def _populate_totals(t_sheet, n_sheet):
    """Create totals from the gathered data in a new worksheet."""

    # Perform some trickery to work out overlapping datasets
    # First get the title row from the networks sheet.
    title_row = n_sheet[1]
    # Find out which columns contain WPAv1 and WPAv2 data
    for cell in title_row:
        if cell.value == 'WPA':
            wpa_col = cell.column
            continue
        if cell.value == 'WPA2':
            wpa2_col = cell.column
    # Create a slices containing the data from both columns
    wpa_cells = n_sheet[f'{wpa_col}']
    wpa2_cells = n_sheet[f'{wpa2_col}']
    # zip those slices to have the cells side by side (excluding the first row)
    wpa_wpa2_cells = zip(wpa_cells[1:], wpa2_cells[1:])
    wpa1_only = 0
    wpa2_only = 0
    wpa_and_wpa2 = 0
    for cells in wpa_wpa2_cells:
        if cells[0].value == 1 and cells[1].value == 0:
            wpa1_only += 1
        elif cells[0].value == 0 and cells[1].value == 1:
            wpa2_only += 1
        elif cells[0].value == 1 and cells[1].value == 1:
            wpa_and_wpa2 += 1

    data = [
        ['Data Set', 'Totals'],
        ['Hidden Networks', '=SUM(Networks[Hidden])'],
        ['Open Networks', '=SUM(Networks[Open])'],
        ['WEP Networks', '=SUM(Networks[WEP])'],
        ['WPAv1 Only', wpa1_only],
        ['WPAv1 And WPAv2', wpa_and_wpa2],
        ['WPAv2 Only', wpa2_only],
        ['Total WPAv1', '=SUM(Networks[WPA])'],
        ['Total WPAv2', '=SUM(Networks[WPA2])'],
        ['WPS Enabled', '=COUNTIF(Networks[WPS], "Configured")'],
        ['TKIP Encryption', '=SUM(Networks[TKIP])'],
        ['AES Encryption', '=SUM(Networks[AES])'],
        ['Total Networks', '=COUNTIF(Networks[BSSID], "*")'],
        # ['Total Client Count', '=SUM(Networks[No. Clients])']
    ]

    for row in data:
        t_sheet.append(row)


def _index_to_column(i, column=''):
    """This took far too long for me to write."""

    # A dictionary of numbers to letters starting at 0, e.g.
    # {0: 'A', 1: 'B' ...}
    num_to_alpha = {k:v for k, v in enumerate(string.ascii_uppercase, 0)}
    # If our index is divisble by 26, we need to get recursive and add
    # additional letters.
    div = i // 26
    if div:
        column = index_to_column(div - 1, column)
    # Combine results in case things got all inception like.
    column = column + num_to_alpha[i % 26]

    return column


def _indices_to_coords(c,r):
    """Take python indices representing column and row - c , r - and
    translate them into excel cordinates.

    For example:
    0, 0 -> A1
    25, 25 -> Z26
    26, 100 -> AA101
    """

    column = _index_to_column(c)
    row = r + 1

    return {'c': column, 'r': row, 'coord': f'{column}{row}'}


if __name__ == "__main__":
      main()
